"""
src/screener/export.py

Sprint 3, Day 17 — generates output/screener_output.xlsx: one sheet per
preset, 20 KPI columns, sorted by composite_score descending, with
green/red conditional fill on every column that has a corresponding
threshold in that preset (green = passes, red = fails).
"""

from __future__ import annotations

from pathlib import Path
from typing import Dict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.analytics.utils import clean

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial")

# 20 KPI columns shown on every preset sheet (plus identifier columns).
KPI_COLUMNS = [
    "composite_score",
    "composite_score_sector_relative",
    "return_on_equity_pct",
    "return_on_capital_employed_pct",
    "return_on_assets_pct",
    "net_profit_margin_pct",
    "operating_profit_margin_pct",
    "debt_to_equity",
    "interest_coverage",
    "asset_turnover",
    "free_cash_flow_cr",
    "fcf_conversion_rate_pct",
    "revenue_cagr_5yr",
    "pat_cagr_5yr",
    "eps_cagr_5yr",
    "pe_ratio",
    "pb_ratio",
    "dividend_yield_pct",
    "market_cap_crore",
    "sales",
]
IDENTIFIER_COLUMNS = ["company_id", "company_name", "broad_sector"]

COLUMN_LABELS = {
    "composite_score": "Composite Score",
    "composite_score_sector_relative": "Composite Score (Sector-Relative)",
    "return_on_equity_pct": "ROE (%)",
    "return_on_capital_employed_pct": "ROCE (%)",
    "return_on_assets_pct": "ROA (%)",
    "net_profit_margin_pct": "NPM (%)",
    "operating_profit_margin_pct": "OPM (%)",
    "debt_to_equity": "D/E",
    "interest_coverage": "Interest Coverage",
    "asset_turnover": "Asset Turnover",
    "free_cash_flow_cr": "FCF (Cr)",
    "fcf_conversion_rate_pct": "FCF Conversion (%)",
    "revenue_cagr_5yr": "Revenue CAGR 5yr (%)",
    "pat_cagr_5yr": "PAT CAGR 5yr (%)",
    "eps_cagr_5yr": "EPS CAGR 5yr (%)",
    "pe_ratio": "P/E",
    "pb_ratio": "P/B",
    "dividend_yield_pct": "Dividend Yield (%)",
    "market_cap_crore": "Market Cap (Cr)",
    "sales": "Sales (Cr)",
    "company_id": "Ticker",
    "company_name": "Company",
    "broad_sector": "Sector",
}

# metric key (from screener_config filterable_metrics) -> KPI column it
# corresponds to, so we know which cell to color for a given preset filter.
_METRIC_TO_KPI_COLUMN = {
    "roe_min": "return_on_equity_pct",
    "de_max": "debt_to_equity",
    "fcf_min": "free_cash_flow_cr",
    "revenue_cagr_5yr_min": "revenue_cagr_5yr",
    "pat_cagr_5yr_min": "pat_cagr_5yr",
    "opm_min": "operating_profit_margin_pct",
    "pe_max": "pe_ratio",
    "pb_max": "pb_ratio",
    "dividend_yield_min": "dividend_yield_pct",
    "icr_min": "interest_coverage",
    "market_cap_min": "market_cap_crore",
    "net_profit_min": None,  # net_profit isn't one of the 20 display columns
    "eps_cagr_5yr_min": "eps_cagr_5yr",
    "asset_turnover_min": "asset_turnover",
    "sales_min": "sales",
}


def _preset_pass_checks(preset_key: str, preset: dict, config: dict) -> Dict[str, callable]:
    """
    Returns {kpi_column: passes_fn(row) -> bool} for every threshold this
    preset defines that maps onto one of the 20 display columns.
    """
    checks = {}
    for f in preset.get("filters", []):
        metric_key = f["metric"]
        threshold = f["threshold"]
        if metric_key not in config["filterable_metrics"]:
            continue
        kpi_col = _METRIC_TO_KPI_COLUMN.get(metric_key)
        if kpi_col is None:
            continue
        spec = config["filterable_metrics"][metric_key]
        direction = spec["direction"]

        def make_check(col, thresh, direction, spec):
            def check(row):
                val = clean(row.get(col))
                if spec.get("debt_free_is_infinity") and row.get("icr_label") == "Debt Free":
                    return True
                if spec.get("skip_financials_sector") and row.get("broad_sector") == "Financials":
                    return True
                if val is None:
                    return False
                return val >= thresh if direction == "min" else val <= thresh

            return check

        checks[kpi_col] = make_check(kpi_col, threshold, direction, spec)

    if preset.get("dividend_payout_max") is not None:
        pass  # dividend_payout_ratio_pct isn't one of the 20 display columns

    if preset.get("de_max_near_zero") is not None:
        thresh = preset["de_max_near_zero"]
        checks["debt_to_equity"] = lambda row, t=thresh: (
            clean(row.get("debt_to_equity")) is not None and clean(row.get("debt_to_equity")) <= t
        )

    if preset.get("fcf_positive_latest_year"):
        checks["free_cash_flow_cr"] = lambda row: (
            clean(row.get("free_cash_flow_cr")) is not None
            and clean(row.get("free_cash_flow_cr")) > 0
        )

    return checks


def _write_sheet(
    wb: Workbook, sheet_name: str, df: pd.DataFrame, preset_key: str, preset: dict, config: dict
):
    ws = wb.create_sheet(title=sheet_name[:31])  # Excel sheet name limit

    columns = IDENTIFIER_COLUMNS + KPI_COLUMNS
    for col_idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=COLUMN_LABELS.get(col, col))
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    pass_checks = _preset_pass_checks(preset_key, preset, config)

    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        for col_idx, col in enumerate(columns, start=1):
            value = row.get(col)
            value = clean(value)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = BODY_FONT
            if col in (
                "return_on_equity_pct",
                "return_on_capital_employed_pct",
                "return_on_assets_pct",
                "net_profit_margin_pct",
                "operating_profit_margin_pct",
                "revenue_cagr_5yr",
                "pat_cagr_5yr",
                "eps_cagr_5yr",
                "fcf_conversion_rate_pct",
                "dividend_yield_pct",
            ):
                cell.number_format = "0.00"
            elif col in (
                "debt_to_equity",
                "interest_coverage",
                "asset_turnover",
                "pe_ratio",
                "pb_ratio",
            ):
                cell.number_format = "0.00"
            elif col in (
                "free_cash_flow_cr",
                "market_cap_crore",
                "sales",
                "composite_score",
                "composite_score_sector_relative",
            ):
                cell.number_format = "#,##0.0"

            if col in pass_checks:
                cell.fill = GREEN_FILL if pass_checks[col](row) else RED_FILL

    for col_idx, col in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(
            12, len(COLUMN_LABELS.get(col, col)) + 2
        )

    ws.freeze_panes = "A2"


def write_screener_output(
    results: Dict[str, pd.DataFrame], config: dict, output_path: Path
) -> Path:
    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet

    for preset_key, df in results.items():
        preset = config["presets"][preset_key]
        _write_sheet(wb, preset["label"], df, preset_key, preset, config)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
