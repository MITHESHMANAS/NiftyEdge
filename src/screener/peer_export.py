"""
src/screener/peer_export.py

Sprint 3, Day 20 — generates output/peer_comparison.xlsx: 11 sheets, one
per peer group. Columns: company_id, company_name, then 20 metric columns
(the same KPI set as screener_output.xlsx) followed by a percentile-rank
column for each of the 10 metrics peer.py ranks. Percentile cells are
colour-coded (green >= 75th pct, yellow 25th-75th, red <= 25th pct), the
benchmark company's row gets a gold/amber background, and a median summary
row sits at the bottom of each sheet.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.analytics.peer import PEER_METRICS
from src.analytics.utils import clean
from src.screener.export import COLUMN_LABELS, IDENTIFIER_COLUMNS, KPI_COLUMNS

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
BENCHMARK_FILL = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
SUMMARY_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial")
BOLD_FONT = Font(name="Arial", bold=True)

PERCENTILE_METRIC_LABELS = list(PEER_METRICS.keys())


def _percentile_fill(pct: float) -> PatternFill:
    if pct >= 0.75:
        return GREEN_FILL
    if pct <= 0.25:
        return RED_FILL
    return YELLOW_FILL


def _write_peer_sheet(
    wb: Workbook,
    group_name: str,
    group_df: pd.DataFrame,
    percentiles: pd.DataFrame,
):
    ws = wb.create_sheet(title=group_name[:31])

    metric_cols = IDENTIFIER_COLUMNS + KPI_COLUMNS
    pct_col_labels = [f"{m} %ile" for m in PERCENTILE_METRIC_LABELS]
    all_cols = metric_cols + pct_col_labels

    for col_idx, label in enumerate(all_cols, start=1):
        header = COLUMN_LABELS.get(label, label)
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    group_df = group_df.sort_values("composite_score", ascending=False).reset_index(drop=True)
    pct_lookup = {
        (row.company_id, row.metric): row.percentile_rank for row in percentiles.itertuples()
    }

    row_idx = 2
    for _, row in group_df.iterrows():
        is_benchmark = bool(row.get("is_benchmark"))
        for col_idx, col in enumerate(metric_cols, start=1):
            value = clean(row.get(col))
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = BOLD_FONT if is_benchmark else BODY_FONT
            if is_benchmark:
                cell.fill = BENCHMARK_FILL

        for i, metric_label in enumerate(PERCENTILE_METRIC_LABELS):
            col_idx = len(metric_cols) + i + 1
            pct = pct_lookup.get((row["company_id"], metric_label))
            cell = ws.cell(row=row_idx, column=col_idx, value=pct)
            cell.font = BOLD_FONT if is_benchmark else BODY_FONT
            cell.number_format = "0.0%"
            if pct is not None:
                cell.fill = _percentile_fill(pct)
            elif is_benchmark:
                cell.fill = BENCHMARK_FILL
        row_idx += 1

    # Summary row: peer group median for each metric column (not the
    # percentile columns — a "median percentile" isn't meaningful).
    summary_row = row_idx
    ws.cell(row=summary_row, column=1, value="Peer Group Median").font = BOLD_FONT
    ws.cell(row=summary_row, column=1).fill = SUMMARY_FILL
    for col_idx in range(2, len(metric_cols) + 1):
        col = metric_cols[col_idx - 1]
        if col in IDENTIFIER_COLUMNS:
            ws.cell(row=summary_row, column=col_idx).fill = SUMMARY_FILL
            continue
        median_val = group_df[col].map(clean).astype(float).median()
        cell = ws.cell(
            row=summary_row, column=col_idx, value=None if pd.isna(median_val) else median_val
        )
        cell.font = BOLD_FONT
        cell.fill = SUMMARY_FILL
    for col_idx in range(len(metric_cols) + 1, len(all_cols) + 1):
        ws.cell(row=summary_row, column=col_idx).fill = SUMMARY_FILL

    for col_idx, label in enumerate(all_cols, start=1):
        header = COLUMN_LABELS.get(label, label)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, len(header) + 2)

    ws.freeze_panes = "A2"


def write_peer_comparison(
    universe: pd.DataFrame, percentiles: pd.DataFrame, output_path: Path
) -> Path:
    """
    `universe` must already have `composite_score` computed (see
    src.screener.composite.compute_composite_scores) and a
    `peer_group_name` / `is_benchmark` column (from
    src.screener.universe.build_screener_universe).
    """
    wb = Workbook()
    wb.remove(wb.active)

    grouped = universe[universe["peer_group_name"].notna()]
    for group_name, group_df in grouped.groupby("peer_group_name"):
        group_pct = percentiles[percentiles["peer_group_name"] == group_name]
        _write_peer_sheet(wb, group_name, group_df, group_pct)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
