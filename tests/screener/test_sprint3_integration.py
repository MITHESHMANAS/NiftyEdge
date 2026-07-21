"""
tests/screener/test_sprint3_integration.py

Sprint 3, Day 21 — integration tests asserting the Sprint 3 Definition of
Done / exit criteria against the real database.
"""

import sqlite3
import sys
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))

from src.analytics.populate_financial_ratios import run_ratio_engine
from src.etl.config import DB_PATH
from src.etl.loader import build_database
from src.screener.run_sprint3 import RADAR_DIR, run_sprint3

ALL_11_PEER_GROUPS = {
    "Automobiles",
    "Consumer Finance",
    "FMCG",
    "IT Services",
    "Life Insurance",
    "Oil & Gas",
    "Pharmaceuticals",
    "Power & Utilities",
    "Private Banks",
    "Public Sector Banks",
    "Steel",
}


@pytest.fixture(scope="module")
def sprint3_result():
    build_database()
    run_ratio_engine()
    return run_sprint3()


class TestPresetExitCriteria:
    def test_all_6_presets_present(self, sprint3_result):
        assert set(sprint3_result["preset_counts"].keys()) == {
            "quality_compounder",
            "value_pick",
            "growth_accelerator",
            "dividend_champion",
            "debt_free_blue_chip",
            "turnaround_watch",
        }

    def test_all_presets_between_5_and_50_companies(self, sprint3_result):
        for preset, n in sprint3_result["preset_counts"].items():
            assert 5 <= n <= 50, f"{preset} returned {n} companies (outside 5-50)"

    def test_quality_compounder_top5_meet_stated_thresholds(self, sprint3_result):
        wb = openpyxl.load_workbook(sprint3_result["screener_output_path"])
        ws = wb["Quality Compounder"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        roe_idx = headers.index("ROE (%)")
        de_idx = headers.index("D/E")
        rows = list(ws.iter_rows(min_row=2, max_row=6, values_only=True))
        for row in rows:
            assert row[roe_idx] > 15
            assert row[de_idx] < 1.0


class TestPeerComparisonExitCriteria:
    def test_peer_comparison_has_exactly_11_sheets(self, sprint3_result):
        wb = openpyxl.load_workbook(sprint3_result["peer_comparison_path"])
        assert set(wb.sheetnames) == ALL_11_PEER_GROUPS
        assert len(wb.sheetnames) == 11

    def test_it_services_highest_roe_has_highest_roe_percentile(self, sprint3_result):
        conn = sqlite3.connect(DB_PATH)
        df = pd.read_sql(
            "SELECT company_id, value, percentile_rank FROM peer_percentiles "
            "WHERE peer_group_name='IT Services' AND metric='ROE' ORDER BY value DESC",
            conn,
        )
        conn.close()
        assert df.iloc[0]["percentile_rank"] == df["percentile_rank"].max()

    def test_fmcg_lowest_de_has_highest_de_percentile(self, sprint3_result):
        conn = sqlite3.connect(DB_PATH)
        df = pd.read_sql(
            "SELECT company_id, value, percentile_rank FROM peer_percentiles "
            "WHERE peer_group_name='FMCG' AND metric='D/E' ORDER BY value ASC",
            conn,
        )
        conn.close()
        assert df.iloc[0]["percentile_rank"] == df["percentile_rank"].max()

    def test_peer_percentiles_cover_all_11_groups(self, sprint3_result):
        assert sprint3_result["peer_result"]["n_groups"] == 11

    def test_no_peer_group_companies_not_silently_dropped(self, sprint3_result):
        # every company either has peer percentile rows OR is listed as
        # having no peer group — none should just vanish.
        conn = sqlite3.connect(DB_PATH)
        all_ids = set(pd.read_sql("SELECT id FROM companies", conn)["id"])
        grouped_ids = set(
            pd.read_sql("SELECT DISTINCT company_id FROM peer_percentiles", conn)["company_id"]
        )
        conn.close()
        no_peer = set(sprint3_result["peer_result"]["no_peer_group_companies"])
        assert all_ids == grouped_ids | no_peer


class TestRadarCharts:
    def test_radar_chart_generated_per_company(self, sprint3_result):
        conn = sqlite3.connect(DB_PATH)
        n_companies = conn.execute("SELECT COUNT(*) FROM companies").fetchone()[0]
        conn.close()
        png_files = list(RADAR_DIR.glob("*_radar.png"))
        assert len(png_files) == n_companies

    def test_radar_chart_filenames_match_ticker_format(self, sprint3_result):
        conn = sqlite3.connect(DB_PATH)
        tcs_exists = conn.execute("SELECT 1 FROM companies WHERE id='TCS'").fetchone()
        conn.close()
        if tcs_exists:
            assert (RADAR_DIR / "TCS_radar.png").exists()


class TestScreenerOutputStructure:
    def test_screener_output_has_6_sheets(self, sprint3_result):
        wb = openpyxl.load_workbook(sprint3_result["screener_output_path"])
        assert len(wb.sheetnames) == 6

    def test_each_sheet_has_20_kpi_plus_identifier_columns(self, sprint3_result):
        from src.screener.export import IDENTIFIER_COLUMNS, KPI_COLUMNS

        wb = openpyxl.load_workbook(sprint3_result["screener_output_path"])
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            n_cols = ws.max_column
            assert n_cols == len(IDENTIFIER_COLUMNS) + len(KPI_COLUMNS)
